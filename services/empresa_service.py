from pathlib import Path
import unicodedata

from openpyxl import Workbook, load_workbook


# ============================================================
# RUTAS Y CONFIGURACIÓN
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent

DATA_DIR = BASE_DIR / "data"

EXCEL_PATH = DATA_DIR / "Control_Electropart.xlsx"

HOJA_EMPRESAS = "Empresas"


# ============================================================
# ESTRUCTURA OFICIAL DE LA HOJA EMPRESAS
# ============================================================

COLUMNAS_EMPRESAS = [
    "ID",
    "NOMBRE",
    "RUC",
    "DIRECCION",
    "EMAIL",
    "TIPO",
    "CREDITO_DIAS",
    "ACTIVO",
]


TIPOS_EMPRESA = {
    "CLIENTE",
    "PROVEEDOR",
    "AMBOS",
}


# ============================================================
# FUNCIONES INTERNAS
# ============================================================

def _normalizar_encabezado(texto):
    """
    Convierte diferentes formas de escribir un encabezado
    a una representación uniforme.

    Ejemplos:

    "Nombre"          -> "NOMBRE"
    "Dirección"       -> "DIRECCION"
    "Crédito días"    -> "CREDITO_DIAS"
    """

    texto = str(texto or "").strip().upper()

    # Separar las letras de sus tildes.
    texto = unicodedata.normalize(
        "NFD",
        texto,
    )

    # Eliminar marcas de acentuación.
    texto = "".join(
        caracter
        for caracter in texto
        if unicodedata.category(caracter) != "Mn"
    )

    texto = texto.replace(" ", "_")

    return texto


def _normalizar_nombre(texto):
    """
    Prepara un nombre para comparar duplicados.

    " Hotel Amazonas " y "HOTEL AMAZONAS"
    se tratarán como el mismo nombre.
    """

    return str(texto or "").strip().casefold()


def _convertir_credito(valor):
    """
    Convierte el crédito almacenado en Excel a un entero.

    Si está vacío o contiene un valor inválido,
    devuelve 0.
    """

    try:
        return int(valor or 0)

    except (TypeError, ValueError):
        return 0


def _convertir_activo(valor):
    """
    Convierte diferentes valores de Excel a True o False.
    """

    if isinstance(valor, bool):
        return valor

    if valor is None:
        return True

    texto = str(valor).strip().lower()

    return texto not in {
        "false",
        "0",
        "no",
        "inactivo",
    }


def _obtener_columnas(hoja):
    """
    Devuelve un diccionario indicando en qué columna
    se encuentra cada encabezado.

    Ejemplo:

    {
        "ID": 1,
        "NOMBRE": 2,
        "RUC": 3,
        "EMAIL": 5
    }
    """

    columnas = {}

    for celda in hoja[1]:

        if celda.value is None:
            continue

        encabezado = _normalizar_encabezado(
            celda.value
        )

        columnas[encabezado] = celda.column

    return columnas


def _asegurar_estructura_empresas(hoja):
    """
    Comprueba que la hoja Empresas tenga todas las columnas
    necesarias.

    Si la hoja antigua solamente tenía:

        ID | Nombre | RUC

    automáticamente la transforma en:

        ID | NOMBRE | RUC | DIRECCION | EMAIL |
        TIPO | CREDITO_DIAS | ACTIVO

    sin eliminar las empresas existentes.

    Retorna True si fue necesario modificar la estructura.
    """

    hubo_cambios = False

    columnas = _obtener_columnas(hoja)

    # --------------------------------------------------------
    # NORMALIZAR ENCABEZADOS EXISTENTES
    # --------------------------------------------------------

    for nombre_columna, indice in columnas.items():

        if nombre_columna in COLUMNAS_EMPRESAS:

            celda = hoja.cell(
                row=1,
                column=indice,
            )

            if celda.value != nombre_columna:

                celda.value = nombre_columna

                hubo_cambios = True

    # Volvemos a obtener las columnas porque algunos
    # encabezados pudieron cambiar.
    columnas = _obtener_columnas(hoja)

    # --------------------------------------------------------
    # AGREGAR COLUMNAS QUE FALTEN
    # --------------------------------------------------------

    for nombre_columna in COLUMNAS_EMPRESAS:

        if nombre_columna not in columnas:

            nueva_columna = hoja.max_column + 1

            hoja.cell(
                row=1,
                column=nueva_columna,
                value=nombre_columna,
            )

            hubo_cambios = True

            columnas[nombre_columna] = nueva_columna

    # --------------------------------------------------------
    # COMPLETAR VALORES PREDETERMINADOS DE REGISTROS ANTIGUOS
    # --------------------------------------------------------

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        nombre = hoja.cell(
            row=fila,
            column=columnas["NOMBRE"],
        ).value

        ruc = hoja.cell(
            row=fila,
            column=columnas["RUC"],
        ).value

        # Si la fila está completamente vacía,
        # no hacemos nada.
        if nombre is None and ruc is None:
            continue

        # Las empresas que ya registramos anteriormente
        # estaban siendo utilizadas como clientes.
        tipo_cell = hoja.cell(
            row=fila,
            column=columnas["TIPO"],
        )

        if tipo_cell.value in (None, ""):

            tipo_cell.value = "CLIENTE"

            hubo_cambios = True

        credito_cell = hoja.cell(
            row=fila,
            column=columnas["CREDITO_DIAS"],
        )

        if credito_cell.value in (None, ""):

            credito_cell.value = 0

            hubo_cambios = True

        activo_cell = hoja.cell(
            row=fila,
            column=columnas["ACTIVO"],
        )

        if activo_cell.value in (None, ""):

            activo_cell.value = True

            hubo_cambios = True

    return hubo_cambios


# ============================================================
# INICIALIZAR ARCHIVO
# ============================================================

def inicializar_archivo():
    """
    Crea el archivo Excel si no existe.

    Si ya existe, actualiza la estructura de Empresas
    sin eliminar los registros existentes.
    """

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    # --------------------------------------------------------
    # EL ARCHIVO TODAVÍA NO EXISTE
    # --------------------------------------------------------

    if not EXCEL_PATH.exists():

        workbook = Workbook()

        hoja = workbook.active

        hoja.title = HOJA_EMPRESAS

        hoja.append(
            COLUMNAS_EMPRESAS
        )

        workbook.save(
            EXCEL_PATH
        )

        workbook.close()

        return

    # --------------------------------------------------------
    # EL ARCHIVO YA EXISTE
    # --------------------------------------------------------

    workbook = load_workbook(
        EXCEL_PATH
    )

    # Si no existe la hoja Empresas,
    # la creamos.
    if HOJA_EMPRESAS not in workbook.sheetnames:

        hoja = workbook.create_sheet(
            HOJA_EMPRESAS
        )

        hoja.append(
            COLUMNAS_EMPRESAS
        )

        workbook.save(
            EXCEL_PATH
        )

        workbook.close()

        return

    # --------------------------------------------------------
    # ACTUALIZAR LA HOJA EXISTENTE
    # --------------------------------------------------------

    hoja = workbook[
        HOJA_EMPRESAS
    ]

    hubo_cambios = (
        _asegurar_estructura_empresas(
            hoja
        )
    )

    # Solo guardamos si realmente tuvimos
    # que modificar algo.
    if hubo_cambios:

        workbook.save(
            EXCEL_PATH
        )

    workbook.close()


# ============================================================
# LISTAR EMPRESAS
# ============================================================

def listar_empresas():
    """
    Devuelve todas las empresas registradas.

    Cada empresa se representa mediante un diccionario.
    """

    inicializar_archivo()

    workbook = load_workbook(
        EXCEL_PATH,
        data_only=True,
    )

    hoja = workbook[
        HOJA_EMPRESAS
    ]

    columnas = _obtener_columnas(
        hoja
    )

    empresas = []

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        empresa_id = hoja.cell(
            fila,
            columnas["ID"],
        ).value

        nombre = hoja.cell(
            fila,
            columnas["NOMBRE"],
        ).value

        ruc = hoja.cell(
            fila,
            columnas["RUC"],
        ).value

        direccion = hoja.cell(
            fila,
            columnas["DIRECCION"],
        ).value

        email = hoja.cell(
            fila,
            columnas["EMAIL"],
        ).value

        tipo = hoja.cell(
            fila,
            columnas["TIPO"],
        ).value

        credito_dias = hoja.cell(
            fila,
            columnas["CREDITO_DIAS"],
        ).value

        activo = hoja.cell(
            fila,
            columnas["ACTIVO"],
        ).value

        # Ignorar filas vacías.
        if nombre is None and ruc is None:
            continue

        empresas.append(
            {
                "id": empresa_id,
                "nombre": str(
                    nombre or ""
                ).strip(),
                "ruc": str(
                    ruc or ""
                ).strip(),
                "direccion": str(
                    direccion or ""
                ).strip(),
                "email": str(
                    email or ""
                ).strip(),
                "tipo": str(
                    tipo or "CLIENTE"
                ).strip().upper(),
                "credito_dias": _convertir_credito(
                    credito_dias
                ),
                "activo": _convertir_activo(
                    activo
                ),
            }
        )

    workbook.close()

    return empresas


# ============================================================
# BUSCAR EMPRESAS
# ============================================================

def buscar_empresas(
    texto,
    campo=None,
    limite=8,
):
    """
    Busca empresas por nombre o RUC.

    Esta función se conserva porque puede ser útil
    para otros módulos aunque Registrar factura
    actualmente utilice una caché en memoria.
    """

    texto = str(
        texto or ""
    ).strip().casefold()

    if texto == "":
        return []

    resultados = []

    for empresa in listar_empresas():

        nombre = empresa[
            "nombre"
        ].casefold()

        ruc = empresa[
            "ruc"
        ].casefold()

        coincide = False

        if campo == "nombre":

            coincide = texto in nombre

        elif campo == "ruc":

            coincide = texto in ruc

        else:

            coincide = (
                texto in nombre
                or texto in ruc
            )

        if coincide:

            resultados.append(
                empresa
            )

        if len(resultados) >= limite:
            break

    return resultados


# ============================================================
# OBTENER EMPRESA POR RUC
# ============================================================

def obtener_empresa_por_ruc(ruc):
    """
    Busca exactamente una empresa por su RUC.
    """

    ruc = str(
        ruc or ""
    ).strip()

    if ruc == "":
        return None

    for empresa in listar_empresas():

        if empresa["ruc"] == ruc:

            return empresa

    return None


# ============================================================
# REGISTRAR EMPRESA
# ============================================================

def registrar_empresa(
    nombre,
    ruc,
    direccion="",
    email="",
    tipo="CLIENTE",
    credito_dias=0,
    activo=True,
):
    """
    Registra una empresa nueva.

    Los parámetros adicionales tienen valores
    predeterminados para mantener compatibilidad
    temporal con la pantalla actual.

    Más adelante empresa_view.py enviará todos
    estos datos explícitamente.
    """

    nombre = str(
        nombre or ""
    ).strip()

    ruc = str(
        ruc or ""
    ).strip()

    direccion = str(
        direccion or ""
    ).strip()

    email = str(
        email or ""
    ).strip()

    tipo = str(
        tipo or "CLIENTE"
    ).strip().upper()
# ============================================================
# OBTENER EMPRESA POR ID
# ============================================================

def obtener_empresa_por_id(empresa_id):
    """
    Busca una empresa utilizando su identificador interno.
    """

    for empresa in listar_empresas():

        if empresa["id"] == empresa_id:
            return empresa

    return None


# ============================================================
# ACTUALIZAR EMPRESA
# ============================================================

def actualizar_empresa(
    empresa_id,
    nombre,
    ruc,
    direccion="",
    email="",
    tipo="CLIENTE",
    credito_dias=0,
    activo=True,
):
    """
    Actualiza los datos de una empresa existente.

    Evita que el nuevo RUC o nombre colisionen
    con otra empresa.
    """

    nombre = str(
        nombre or ""
    ).strip()

    ruc = str(
        ruc or ""
    ).strip()

    direccion = str(
        direccion or ""
    ).strip()

    email = str(
        email or ""
    ).strip()

    tipo = str(
        tipo or "CLIENTE"
    ).strip().upper()

    # --------------------------------------------------------
    # VALIDACIONES INTERNAS
    # --------------------------------------------------------

    if nombre == "":
        return (
            False,
            "El nombre de la empresa es obligatorio.",
        )

    if ruc == "":
        return (
            False,
            "El RUC es obligatorio.",
        )

    if tipo not in TIPOS_EMPRESA:
        return (
            False,
            "El tipo de empresa no es válido.",
        )

    try:

        credito_dias = int(
            credito_dias or 0
        )

    except (TypeError, ValueError):

        return (
            False,
            "Los días de crédito deben ser un número entero.",
        )

    if credito_dias < 0:
        return (
            False,
            "Los días de crédito no pueden ser negativos.",
        )

    # --------------------------------------------------------
    # ABRIR EXCEL
    # --------------------------------------------------------

    inicializar_archivo()

    workbook = load_workbook(
        EXCEL_PATH
    )

    hoja = workbook[
        HOJA_EMPRESAS
    ]

    columnas = _obtener_columnas(
        hoja
    )

    fila_objetivo = None

    nombre_normalizado = _normalizar_nombre(
        nombre
    )

    # --------------------------------------------------------
    # BUSCAR EMPRESA Y VALIDAR DUPLICADOS
    # --------------------------------------------------------

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        id_existente = hoja.cell(
            fila,
            columnas["ID"],
        ).value

        nombre_existente = str(
            hoja.cell(
                fila,
                columnas["NOMBRE"],
            ).value
            or ""
        ).strip()

        ruc_existente = str(
            hoja.cell(
                fila,
                columnas["RUC"],
            ).value
            or ""
        ).strip()

        # Encontramos la fila que vamos a modificar.
        if id_existente == empresa_id:

            fila_objetivo = fila

            continue

        # RUC perteneciente a OTRA empresa.
        if (
            ruc_existente != ""
            and ruc_existente == ruc
        ):

            workbook.close()

            return (
                False,
                "Otra empresa ya utiliza este RUC.",
            )

        # Nombre perteneciente a OTRA empresa.
        if (
            nombre_existente != ""
            and _normalizar_nombre(
                nombre_existente
            ) == nombre_normalizado
        ):

            workbook.close()

            return (
                False,
                "Otra empresa ya utiliza este nombre.",
            )

    # --------------------------------------------------------
    # EMPRESA NO ENCONTRADA
    # --------------------------------------------------------

    if fila_objetivo is None:

        workbook.close()

        return (
            False,
            "No se encontró la empresa que desea modificar.",
        )

    # --------------------------------------------------------
    # ACTUALIZAR CELDAS
    # --------------------------------------------------------

    hoja.cell(
        fila_objetivo,
        columnas["NOMBRE"],
        nombre,
    )

    hoja.cell(
        fila_objetivo,
        columnas["RUC"],
        ruc,
    )

    hoja.cell(
        fila_objetivo,
        columnas["DIRECCION"],
        direccion,
    )

    hoja.cell(
        fila_objetivo,
        columnas["EMAIL"],
        email,
    )

    hoja.cell(
        fila_objetivo,
        columnas["TIPO"],
        tipo,
    )

    hoja.cell(
        fila_objetivo,
        columnas["CREDITO_DIAS"],
        credito_dias,
    )

    hoja.cell(
        fila_objetivo,
        columnas["ACTIVO"],
        bool(activo),
    )

    workbook.save(
        EXCEL_PATH
    )

    workbook.close()

    return (
        True,
        "Empresa actualizada correctamente.",
    )
    # --------------------------------------------------------
    # VALIDACIONES INTERNAS BÁSICAS
    # --------------------------------------------------------

    if nombre == "":

        return (
            False,
            "El nombre de la empresa es obligatorio.",
        )

    if ruc == "":

        return (
            False,
            "El RUC es obligatorio.",
        )

    if tipo not in TIPOS_EMPRESA:

        return (
            False,
            "El tipo de empresa no es válido.",
        )

    try:

        credito_dias = int(
            credito_dias or 0
        )

    except (TypeError, ValueError):

        return (
            False,
            "Los días de crédito deben ser un número entero.",
        )

    if credito_dias < 0:

        return (
            False,
            "Los días de crédito no pueden ser negativos.",
        )

    inicializar_archivo()

    workbook = load_workbook(
        EXCEL_PATH
    )

    hoja = workbook[
        HOJA_EMPRESAS
    ]

    columnas = _obtener_columnas(
        hoja
    )

    # --------------------------------------------------------
    # COMPROBAR DUPLICADOS
    # --------------------------------------------------------

    nombre_normalizado = (
        _normalizar_nombre(
            nombre
        )
    )

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        nombre_existente = hoja.cell(
            fila,
            columnas["NOMBRE"],
        ).value

        ruc_existente = hoja.cell(
            fila,
            columnas["RUC"],
        ).value

        nombre_existente = str(
            nombre_existente or ""
        ).strip()

        ruc_existente = str(
            ruc_existente or ""
        ).strip()

        if (
            ruc_existente != ""
            and ruc_existente == ruc
        ):

            workbook.close()

            return (
                False,
                "Ya existe una empresa registrada con este RUC.",
            )

        if (
            nombre_existente != ""
            and _normalizar_nombre(
                nombre_existente
            ) == nombre_normalizado
        ):

            workbook.close()

            return (
                False,
                "Ya existe una empresa registrada con este nombre.",
            )

    # --------------------------------------------------------
    # GENERAR NUEVO ID
    # --------------------------------------------------------

    ids = []

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        valor_id = hoja.cell(
            fila,
            columnas["ID"],
        ).value

        if isinstance(
            valor_id,
            int,
        ):

            ids.append(
                valor_id
            )

    nuevo_id = max(
        ids,
        default=0,
    ) + 1

    # --------------------------------------------------------
    # CREAR NUEVA FILA
    # --------------------------------------------------------

    nueva_fila = hoja.max_row + 1

    hoja.cell(
        nueva_fila,
        columnas["ID"],
        nuevo_id,
    )

    hoja.cell(
        nueva_fila,
        columnas["NOMBRE"],
        nombre,
    )

    hoja.cell(
        nueva_fila,
        columnas["RUC"],
        ruc,
    )

    hoja.cell(
        nueva_fila,
        columnas["DIRECCION"],
        direccion,
    )

    hoja.cell(
        nueva_fila,
        columnas["EMAIL"],
        email,
    )

    hoja.cell(
        nueva_fila,
        columnas["TIPO"],
        tipo,
    )

    hoja.cell(
        nueva_fila,
        columnas["CREDITO_DIAS"],
        credito_dias,
    )

    hoja.cell(
        nueva_fila,
        columnas["ACTIVO"],
        bool(activo),
    )

    # --------------------------------------------------------
    # GUARDAR
    # --------------------------------------------------------

    workbook.save(
        EXCEL_PATH
    )

    workbook.close()

    return (
        True,
        "Empresa registrada correctamente.",
    )
def actualizar_empresa(
    empresa_id,
    nombre,
    ruc,
    direccion="",
    email="",
    tipo="CLIENTE",
    credito_dias=0,
    activo=True,
):
    """
    Actualiza una empresa existente utilizando
    su ID interno.

    El RUC continúa siendo único.
    """

    nombre = str(nombre or "").strip()
    ruc = str(ruc or "").strip()
    direccion = str(direccion or "").strip()
    email = str(email or "").strip()
    tipo = str(tipo or "CLIENTE").strip().upper()

    try:
        credito_dias = int(credito_dias or 0)

    except (TypeError, ValueError):
        return (
            False,
            "Los días de crédito deben ser un número entero.",
        )

    inicializar_archivo()

    workbook = load_workbook(EXCEL_PATH)

    hoja = workbook[HOJA_EMPRESAS]

    columnas = _obtener_columnas(hoja)

    fila_objetivo = None

    nombre_normalizado = _normalizar_nombre(nombre)

    # --------------------------------------------------------
    # BUSCAR EMPRESA Y COMPROBAR DUPLICADOS
    # --------------------------------------------------------

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        id_existente = hoja.cell(
            fila,
            columnas["ID"],
        ).value

        nombre_existente = str(
            hoja.cell(
                fila,
                columnas["NOMBRE"],
            ).value
            or ""
        ).strip()

        ruc_existente = str(
            hoja.cell(
                fila,
                columnas["RUC"],
            ).value
            or ""
        ).strip()

        # Esta es la propia empresa.
        if id_existente == empresa_id:

            fila_objetivo = fila

            continue

        # Otra empresa tiene el mismo RUC.
        if (
            ruc_existente != ""
            and ruc_existente == ruc
        ):

            workbook.close()

            return (
                False,
                "Otra empresa ya utiliza este RUC.",
            )

        # Otra empresa tiene exactamente el mismo nombre.
        if (
            nombre_existente != ""
            and _normalizar_nombre(
                nombre_existente
            ) == nombre_normalizado
        ):

            workbook.close()

            return (
                False,
                "Otra empresa ya utiliza este nombre.",
            )

    if fila_objetivo is None:

        workbook.close()

        return (
            False,
            "No se encontró la empresa.",
        )

    # --------------------------------------------------------
    # ACTUALIZAR
    # --------------------------------------------------------

    hoja.cell(
        fila_objetivo,
        columnas["NOMBRE"],
        nombre,
    )

    hoja.cell(
        fila_objetivo,
        columnas["RUC"],
        ruc,
    )

    hoja.cell(
        fila_objetivo,
        columnas["DIRECCION"],
        direccion,
    )

    hoja.cell(
        fila_objetivo,
        columnas["EMAIL"],
        email,
    )

    hoja.cell(
        fila_objetivo,
        columnas["TIPO"],
        tipo,
    )

    hoja.cell(
        fila_objetivo,
        columnas["CREDITO_DIAS"],
        credito_dias,
    )

    hoja.cell(
        fila_objetivo,
        columnas["ACTIVO"],
        bool(activo),
    )

    workbook.save(EXCEL_PATH)

    workbook.close()

    return (
        True,
        "Empresa actualizada correctamente.",
    )