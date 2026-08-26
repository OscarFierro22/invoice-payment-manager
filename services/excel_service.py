from pathlib import Path
import unicodedata

from openpyxl import Workbook, load_workbook


# ============================================================
# UBICACIÓN DEL ARCHIVO
# ============================================================

BASE_DIR = Path(__file__).resolve().parent.parent

DATA_DIR = BASE_DIR / "data"

EXCEL_PATH = DATA_DIR / "Control_Electropart.xlsx"


# ============================================================
# ESTRUCTURA OFICIAL DEL SISTEMA
# ============================================================

ESTRUCTURA_HOJAS = {

    # --------------------------------------------------------
    # EMPRESAS
    # --------------------------------------------------------

    "Empresas": [
        "ID",
        "NOMBRE",
        "RUC",
        "DIRECCION",
        "EMAIL",
        "TIPO",
        "CREDITO_DIAS",
        "ACTIVO",
    ],
    # --------------------------------------------------------
    # PRODUCTOS
    # --------------------------------------------------------

    "Productos": [
        "ID_PRODUCTO",
        "NOMBRE",
        "CATEGORIA",
        "MARCA",
        "MODELO",
        "CODIGO_INTERNO",
        "UNIDAD",
        "DESCRIPCION",
        "ACTIVO",
        "FECHA_REGISTRO",
    ],
    # --------------------------------------------------------
    # COMPRAS
    # --------------------------------------------------------

    "Compras": [
        "ID_COMPRA",
        "FECHA_COMPRA",
        "ID_PROVEEDOR",
        "CANTIDAD",
        "DESCRIPCION",
        "COSTO_UNITARIO",
        "SUBTOTAL_COMPRA",
        "IVA_PORCENTAJE",
        "IVA_VALOR",
        "TOTAL_COMPRA",
        "RECARGO_PORCENTAJE",
        "PRECIO_FACTURAR",
        "DESTINO_TIPO",
        "ID_CLIENTE",
        "DESTINO_TEXTO",
        "ESTADO_FACTURACION",
        "ID_FACTURA",
        "FECHA_REGISTRO",
        "ORIGEN_HOJA",
        "ORIGEN_FILA",
    ],

    # --------------------------------------------------------
    # FACTURAS - CABECERA
    # --------------------------------------------------------

    "Facturas": [
    "ID_FACTURA",
    "ID_EMPRESA",
    "EMPRESA",
    "RUC",
    "NUMERO_FACTURA",
    "FECHA_EMISION",
    "CREDITO_DIAS",
    "FECHA_VENCIMIENTO",
    "SUBTOTAL",
    "IVA_PORCENTAJE",
    "IVA",
    "TOTAL",
    "ESTADO_DOCUMENTO",
    "FECHA_REGISTRO",
    "ORIGEN_HOJA",
    "ORIGEN_FILA",
],

    # --------------------------------------------------------
    # DETALLE DE FACTURA
    # --------------------------------------------------------

    "DetalleFacturas": [
        "ID_DETALLE",
        "ID_FACTURA",
        "TIPO_ITEM",
        "DESCRIPCION",
        "CANTIDAD",
        "PRECIO_UNITARIO",
        "SUBTOTAL",
        "ID_COMPRA",
        "FECHA_REGISTRO",
    ],

    # --------------------------------------------------------
    # RETENCIONES - CABECERA
    # --------------------------------------------------------

    "Retenciones": [
        "ID_RETENCION",
        "ID_FACTURA",
        "NUMERO_COMPROBANTE",
        "FECHA_RETENCION",
        "TOTAL_RETENIDO",
        "ESTADO",
        "OBSERVACION",
        "FECHA_REGISTRO",
        "ORIGEN_HOJA",
        "ORIGEN_FILA",
    ],

    # --------------------------------------------------------
    # DETALLE DE RETENCIONES
    # --------------------------------------------------------

    "DetalleRetenciones": [
        "ID_DETALLE_RETENCION",
        "ID_RETENCION",
        "TIPO",
        "BASE_IMPONIBLE",
        "PORCENTAJE",
        "VALOR",
    ],

    # --------------------------------------------------------
    # PAGOS
    # --------------------------------------------------------

    "Pagos": [
        "ID_PAGO",
        "ID_FACTURA",
        "FECHA_PAGO",
        "VALOR",
        "METODO",
        "REFERENCIA",
        "OBSERVACION",
        "FECHA_REGISTRO",
        "ORIGEN_HOJA",
        "ORIGEN_FILA",
    ],

    # --------------------------------------------------------
    # CONFIGURACIÓN
    # --------------------------------------------------------

    "Configuracion": [
        "CLAVE",
        "VALOR",
        "DESCRIPCION",
    ],

    # --------------------------------------------------------
    # AUDITORÍA
    # --------------------------------------------------------

    "Auditoria": [
        "ID_AUDITORIA",
        "FECHA_HORA",
        "ACCION",
        "ENTIDAD",
        "ID_ENTIDAD",
        "DETALLE",
    ],
}


# ============================================================
# NORMALIZAR ENCABEZADOS
# ============================================================

def normalizar_encabezado(texto):
    """
    Convierte encabezados a un formato comparable.

    Ejemplos:

        "Dirección"       -> "DIRECCION"
        "Crédito días"    -> "CREDITO_DIAS"
        "Fecha Emisión"   -> "FECHA_EMISION"
    """

    texto = str(texto or "").strip().upper()

    texto = unicodedata.normalize(
        "NFD",
        texto,
    )

    texto = "".join(
        caracter
        for caracter in texto
        if unicodedata.category(caracter) != "Mn"
    )

    texto = texto.replace(
        " ",
        "_",
    )

    return texto


# ============================================================
# OBTENER COLUMNAS DE UNA HOJA
# ============================================================

def obtener_columnas(hoja):
    """
    Devuelve un diccionario con el nombre de cada columna
    y su posición dentro de Excel.

    Ejemplo:

    {
        "ID_FACTURA": 1,
        "ID_EMPRESA": 2,
        "EMPRESA": 3,
    }
    """

    columnas = {}

    for numero_columna in range(
        1,
        hoja.max_column + 1,
    ):

        valor = hoja.cell(
            1,
            numero_columna,
        ).value

        if valor is None:
            continue

        nombre = normalizar_encabezado(
            valor
        )

        columnas[nombre] = numero_columna

    return columnas


# ============================================================
# ASEGURAR COLUMNAS
# ============================================================

def asegurar_columnas(
    hoja,
    columnas_requeridas,
):
    """
    Comprueba que una hoja tenga todas las columnas
    requeridas por el sistema.

    Si falta alguna columna, la agrega al final.

    IMPORTANTE:
    - no elimina columnas;
    - no elimina datos;
    - no modifica registros existentes.
    """

    hubo_cambios = False

    columnas_actuales = obtener_columnas(
        hoja
    )

    for columna in columnas_requeridas:

        nombre_normalizado = (
            normalizar_encabezado(
                columna
            )
        )

        if (
            nombre_normalizado
            not in columnas_actuales
        ):

            nueva_columna = (
                hoja.max_column + 1
            )

            hoja.cell(
                1,
                nueva_columna,
                columna,
            )

            columnas_actuales[
                nombre_normalizado
            ] = nueva_columna

            hubo_cambios = True

    return hubo_cambios


# ============================================================
# INICIALIZAR ESTRUCTURA
# ============================================================

def inicializar_estructura():
    """
    Garantiza que Control_Electropart.xlsx tenga todas
    las hojas necesarias para el sistema.

    Esta función funciona como una migración segura:

    - no elimina registros;
    - no elimina hojas;
    - no elimina columnas;
    - solamente crea hojas o columnas faltantes.
    """

    # --------------------------------------------------------
    # ASEGURAR CARPETA DATA
    # --------------------------------------------------------

    DATA_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    # ========================================================
    # CASO 1:
    # EL ARCHIVO TODAVÍA NO EXISTE
    # ========================================================

    if not EXCEL_PATH.exists():

        workbook = Workbook()

        # ----------------------------------------------------
        # CREAR HOJA EMPRESAS
        # ----------------------------------------------------

        hoja_inicial = workbook.active

        hoja_inicial.title = "Empresas"

        for indice, columna in enumerate(
            ESTRUCTURA_HOJAS["Empresas"],
            start=1,
        ):

            hoja_inicial.cell(
                1,
                indice,
                columna,
            )
        
        # ----------------------------------------------------
        # CREAR LAS DEMÁS HOJAS
        # ----------------------------------------------------

        for nombre_hoja, columnas in (
            ESTRUCTURA_HOJAS.items()
        ):

            # Empresas ya fue creada anteriormente.
            if nombre_hoja == "Empresas":
                continue

            hoja = workbook.create_sheet(
                nombre_hoja
            )

            for indice, columna in enumerate(
                columnas,
                start=1,
            ):

                hoja.cell(
                    1,
                    indice,
                    columna,
                )

        # ----------------------------------------------------
        # GUARDAR ARCHIVO NUEVO
        # ----------------------------------------------------

        workbook.save(
            EXCEL_PATH
        )

        workbook.close()

        return

    # ========================================================
    # CASO 2:
    # EL ARCHIVO YA EXISTE
    # ========================================================

    workbook = load_workbook(
        EXCEL_PATH
    )

    hubo_cambios = False

    # --------------------------------------------------------
    # RECORRER TODA LA ESTRUCTURA OFICIAL
    # --------------------------------------------------------

    for nombre_hoja, columnas in (
        ESTRUCTURA_HOJAS.items()
    ):

        # ====================================================
        # LA HOJA NO EXISTE
        # ====================================================

        if nombre_hoja not in workbook.sheetnames:

            hoja = workbook.create_sheet(
                nombre_hoja
            )

            for indice, columna in enumerate(
                columnas,
                start=1,
            ):

                hoja.cell(
                    1,
                    indice,
                    columna,
                )

            hubo_cambios = True

            continue

        # ====================================================
        # LA HOJA YA EXISTE
        # ====================================================

        hoja = workbook[
            nombre_hoja
        ]

        # Verificar si le falta alguna columna.
        if asegurar_columnas(
            hoja,
            columnas,
        ):

            hubo_cambios = True

    # --------------------------------------------------------
    # GUARDAR SOLAMENTE SI HUBO CAMBIOS
    # --------------------------------------------------------

    if hubo_cambios:

        workbook.save(
            EXCEL_PATH
        )

    workbook.close()