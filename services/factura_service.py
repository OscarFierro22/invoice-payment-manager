from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import os

from openpyxl import load_workbook

from services.excel_service import (
    EXCEL_PATH,
    inicializar_estructura,
    obtener_columnas,
)


# ============================================================
# HOJAS
# ============================================================

HOJA_FACTURAS = "Facturas"

HOJA_DETALLE_FACTURAS = "DetalleFacturas"


# ============================================================
# TIPOS DE ÍTEMS QUE PUEDE FACTURAR ELECTROPART
# ============================================================

TIPOS_ITEM = {
    "EQUIPO",
    "REPUESTO",
    "MATERIAL",
    "SERVICIO",
    "MANTENIMIENTO",
    "MANO_DE_OBRA",
    "OTRO",
}


# ============================================================
# CONSTANTES
# ============================================================

DOS_DECIMALES = Decimal("0.01")


# ============================================================
# CONVERTIR A DECIMAL
# ============================================================

def _convertir_decimal(
    valor,
    nombre_campo,
):
    """
    Convierte un valor a Decimal.

    Decimal es preferible a float cuando trabajamos
    con dinero.

    Ejemplo:

        "25.50"
            ↓
        Decimal("25.50")
    """

    texto = str(
        valor if valor is not None else ""
    ).strip()

    texto = texto.replace(
        ",",
        ".",
    )

    if texto == "":
        raise ValueError(
            f"El campo '{nombre_campo}' es obligatorio."
        )

    try:

        numero = Decimal(
            texto
        )

    except InvalidOperation:

        raise ValueError(
            f"El campo '{nombre_campo}' debe ser numérico."
        )

    return numero


# ============================================================
# REDONDEAR DINERO
# ============================================================

def _redondear_moneda(valor):
    """
    Redondea valores monetarios a dos decimales.

    Ejemplo:

        10.236
            ↓
        10.24
    """

    return Decimal(valor).quantize(
        DOS_DECIMALES,
        rounding=ROUND_HALF_UP,
    )


# ============================================================
# CONVERTIR FECHA
# ============================================================

def _convertir_fecha(valor):
    """
    Acepta:

    - datetime
    - date
    - texto dd/mm/aaaa

    Devuelve datetime.
    """

    if isinstance(
        valor,
        datetime,
    ):
        return valor

    if isinstance(
        valor,
        date,
    ):
        return datetime(
            valor.year,
            valor.month,
            valor.day,
        )

    texto = str(
        valor or ""
    ).strip()

    if texto == "":
        raise ValueError(
            "La fecha de emisión es obligatoria."
        )

    try:

        return datetime.strptime(
            texto,
            "%d/%m/%Y",
        )

    except ValueError:

        raise ValueError(
            "La fecha debe tener formato dd/mm/aaaa y ser válida."
        )


# ============================================================
# NORMALIZAR TIPO DE ÍTEM
# ============================================================

def _normalizar_tipo_item(tipo):
    """
    Ejemplos:

        "mano de obra"
            ↓
        "MANO_DE_OBRA"

        "Mantenimiento"
            ↓
        "MANTENIMIENTO"
    """

    tipo = str(
        tipo or ""
    ).strip().upper()

    tipo = tipo.replace(
        " ",
        "_",
    )

    return tipo


# ============================================================
# OBTENER SIGUIENTE ID
# ============================================================

def _obtener_siguiente_id(
    hoja,
    columna_id,
):
    """
    Obtiene el siguiente ID disponible.

    Ejemplo:

        1
        2
        3

        siguiente → 4
    """

    mayor_id = 0

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        valor = hoja.cell(
            fila,
            columna_id,
        ).value

        if valor is None:
            continue

        try:

            valor = int(
                valor
            )

        except (TypeError, ValueError):
            continue

        if valor > mayor_id:
            mayor_id = valor

    return mayor_id + 1


# ============================================================
# ESCRIBIR FILA USANDO NOMBRES DE COLUMNAS
# ============================================================

def _escribir_fila(
    hoja,
    columnas,
    datos,
):
    """
    Escribe una fila sin depender del orden físico
    de las columnas de Excel.

    Ejemplo:

        {
            "ID_FACTURA": 10,
            "EMPRESA": "Sheraton",
            ...
        }
    """

    nueva_fila = (
        hoja.max_row + 1
    )

    for nombre_columna, valor in (
        datos.items()
    ):

        numero_columna = columnas.get(
            nombre_columna
        )

        if numero_columna is None:

            raise ValueError(
                f"No existe la columna "
                f"'{nombre_columna}' en Excel."
            )

        hoja.cell(
            nueva_fila,
            numero_columna,
            valor,
        )

    return nueva_fila


# ============================================================
# COMPROBAR NÚMERO DE FACTURA
# ============================================================

def _numero_factura_existe(
    hoja_facturas,
    columnas_facturas,
    numero_factura,
):
    """
    Una factura emitida por ELECTROPART no debe
    registrarse dos veces con el mismo número.
    """

    numero_buscado = str(
        numero_factura
    ).strip().casefold()

    for fila in range(
        2,
        hoja_facturas.max_row + 1,
    ):

        numero_existente = str(
            hoja_facturas.cell(
                fila,
                columnas_facturas[
                    "NUMERO_FACTURA"
                ],
            ).value
            or ""
        ).strip().casefold()

        if (
            numero_existente
            == numero_buscado
        ):

            return True

    return False


# ============================================================
# CALCULAR DETALLES Y TOTALES
# ============================================================

def calcular_totales(
    detalles,
    iva_porcentaje,
):
    """
    Valida los ítems y calcula:

    - subtotal;
    - IVA;
    - total.

    NO modifica Excel.

    Esto permite utilizar esta función directamente
    desde la interfaz antes de guardar la factura.
    """

    if not detalles:

        raise ValueError(
            "La factura debe contener al menos un detalle."
        )

    iva_porcentaje = _convertir_decimal(
        iva_porcentaje,
        "IVA",
    )

    if iva_porcentaje < 0:

        raise ValueError(
            "El porcentaje de IVA no puede ser negativo."
        )

    detalles_normalizados = []

    subtotal_factura = Decimal(
        "0"
    )

    for numero, detalle in enumerate(
        detalles,
        start=1,
    ):

        # ----------------------------------------------------
        # TIPO
        # ----------------------------------------------------

        tipo_item = _normalizar_tipo_item(
            detalle.get(
                "tipo_item"
            )
        )

        if tipo_item not in TIPOS_ITEM:

            raise ValueError(
                f"El tipo del detalle #{numero} "
                f"no es válido."
            )

        # ----------------------------------------------------
        # DESCRIPCIÓN
        # ----------------------------------------------------

        descripcion = str(
            detalle.get(
                "descripcion",
                "",
            )
        ).strip()

        if descripcion == "":

            raise ValueError(
                f"La descripción del detalle "
                f"#{numero} es obligatoria."
            )

        # ----------------------------------------------------
        # CANTIDAD
        # ----------------------------------------------------

        cantidad = _convertir_decimal(
            detalle.get(
                "cantidad"
            ),
            f"Cantidad del detalle #{numero}",
        )

        if cantidad <= 0:

            raise ValueError(
                f"La cantidad del detalle "
                f"#{numero} debe ser mayor que cero."
            )

        # ----------------------------------------------------
        # PRECIO UNITARIO
        # ----------------------------------------------------

        precio_unitario = _convertir_decimal(
            detalle.get(
                "precio_unitario"
            ),
            f"Precio del detalle #{numero}",
        )

        if precio_unitario < 0:

            raise ValueError(
                f"El precio del detalle "
                f"#{numero} no puede ser negativo."
            )

        precio_unitario = (
            _redondear_moneda(
                precio_unitario
            )
        )

        # ----------------------------------------------------
        # SUBTOTAL DEL ÍTEM
        # ----------------------------------------------------

        subtotal_item = (
            cantidad
            * precio_unitario
        )

        subtotal_item = (
            _redondear_moneda(
                subtotal_item
            )
        )

        subtotal_factura += (
            subtotal_item
        )

        # ----------------------------------------------------
        # COMPRA RELACIONADA
        # ----------------------------------------------------

        id_compra = detalle.get(
            "id_compra"
        )

        # ----------------------------------------------------
        # GUARDAR DETALLE NORMALIZADO
        # ----------------------------------------------------

        detalles_normalizados.append(
            {
                "tipo_item": tipo_item,
                "descripcion": descripcion,
                "cantidad": cantidad,
                "precio_unitario": precio_unitario,
                "subtotal": subtotal_item,
                "id_compra": id_compra,
            }
        )

    # ========================================================
    # SUBTOTAL
    # ========================================================

    subtotal_factura = (
        _redondear_moneda(
            subtotal_factura
        )
    )

    # ========================================================
    # IVA
    # ========================================================

    iva_valor = (
        subtotal_factura
        * iva_porcentaje
        / Decimal("100")
    )

    iva_valor = (
        _redondear_moneda(
            iva_valor
        )
    )

    # ========================================================
    # TOTAL
    # ========================================================

    total = (
        subtotal_factura
        + iva_valor
    )

    total = _redondear_moneda(
        total
    )

    return {
        "detalles": detalles_normalizados,
        "subtotal": subtotal_factura,
        "iva_porcentaje": iva_porcentaje,
        "iva": iva_valor,
        "total": total,
    }


# ============================================================
# REGISTRAR FACTURA
# ============================================================

def registrar_factura(
    id_empresa,
    empresa,
    ruc,
    numero_factura,
    fecha_emision,
    credito_dias,
    iva_porcentaje,
    detalles,
):
    """
    Registra:

        1 factura
            +
        todos sus detalles

    El archivo se guarda una sola vez al final.
    """

    # ========================================================
    # VALIDACIONES GENERALES
    # ========================================================

    empresa = str(
        empresa or ""
    ).strip()

    ruc = str(
        ruc or ""
    ).strip()

    numero_factura = str(
        numero_factura or ""
    ).strip()

    if id_empresa is None:

        return (
            False,
            "Debe seleccionar una empresa válida.",
        )

    if empresa == "":

        return (
            False,
            "El nombre de la empresa es obligatorio.",
        )

    if ruc == "":

        return (
            False,
            "El RUC es obligatorio.",
        )

    if numero_factura == "":

        return (
            False,
            "El número de factura es obligatorio.",
        )

    # ========================================================
    # FECHA
    # ========================================================

    try:

        fecha_emision = (
            _convertir_fecha(
                fecha_emision
            )
        )

    except ValueError as error:

        return (
            False,
            str(error),
        )

    # ========================================================
    # CRÉDITO
    # ========================================================

    try:

        credito_dias = int(
            credito_dias or 0
        )

    except (TypeError, ValueError):

        return (
            False,
            "Los días de crédito deben ser un número entero.",
        )

    if credito_dias < 0:

        return (
            False,
            "Los días de crédito no pueden ser negativos.",
        )

    fecha_vencimiento = (
        fecha_emision
        + timedelta(
            days=credito_dias
        )
    )

    # ========================================================
    # CALCULAR FACTURA
    # ========================================================

    try:

        calculos = calcular_totales(
            detalles=detalles,
            iva_porcentaje=iva_porcentaje,
        )

    except ValueError as error:

        return (
            False,
            str(error),
        )

    # ========================================================
    # PREPARAR EXCEL
    # ========================================================

    inicializar_estructura()

    try:

        workbook = load_workbook(
            EXCEL_PATH
        )

    except PermissionError:

        return (
            False,
            "No se puede abrir el archivo Excel. "
            "Compruebe que esté cerrado.",
        )

    hoja_facturas = workbook[
        HOJA_FACTURAS
    ]

    hoja_detalles = workbook[
        HOJA_DETALLE_FACTURAS
    ]

    columnas_facturas = (
        obtener_columnas(
            hoja_facturas
        )
    )

    columnas_detalles = (
        obtener_columnas(
            hoja_detalles
        )
    )

    # ========================================================
    # EVITAR FACTURAS DUPLICADAS
    # ========================================================

    if _numero_factura_existe(
        hoja_facturas,
        columnas_facturas,
        numero_factura,
    ):

        workbook.close()

        return (
            False,
            "Ya existe una factura con ese número.",
        )

    # ========================================================
    # GENERAR ID DE FACTURA
    # ========================================================

    id_factura = (
        _obtener_siguiente_id(
            hoja_facturas,
            columnas_facturas[
                "ID_FACTURA"
            ],
        )
    )

    fecha_registro = (
        datetime.now()
    )

    # ========================================================
    # ESCRIBIR CABECERA
    # ========================================================

    try:

        _escribir_fila(
            hoja_facturas,
            columnas_facturas,
            {
                "ID_FACTURA": id_factura,
                "ID_EMPRESA": id_empresa,
                "EMPRESA": empresa,
                "RUC": ruc,
                "NUMERO_FACTURA": numero_factura,
                "FECHA_EMISION": fecha_emision,
                "CREDITO_DIAS": credito_dias,
                "FECHA_VENCIMIENTO": fecha_vencimiento,
                "SUBTOTAL": float(
                    calculos["subtotal"]
                ),
                "IVA_PORCENTAJE": float(
                    calculos[
                        "iva_porcentaje"
                    ]
                ),
                "IVA": float(
                    calculos["iva"]
                ),
                "TOTAL": float(
                    calculos["total"]
                ),
                "ESTADO_DOCUMENTO": "EMITIDA",
                "FECHA_REGISTRO": fecha_registro,
                "ORIGEN_HOJA": "",
                "ORIGEN_FILA": "",
            },
        )

        # ====================================================
        # GENERAR IDS DE DETALLES
        # ====================================================

        siguiente_id_detalle = (
            _obtener_siguiente_id(
                hoja_detalles,
                columnas_detalles[
                    "ID_DETALLE"
                ],
            )
        )

        # ====================================================
        # ESCRIBIR DETALLES
        # ====================================================

        for detalle in calculos[
            "detalles"
        ]:

            _escribir_fila(
                hoja_detalles,
                columnas_detalles,
                {
                    "ID_DETALLE":
                        siguiente_id_detalle,

                    "ID_FACTURA":
                        id_factura,

                    "TIPO_ITEM":
                        detalle[
                            "tipo_item"
                        ],

                    "DESCRIPCION":
                        detalle[
                            "descripcion"
                        ],

                    "CANTIDAD":
                        float(
                            detalle[
                                "cantidad"
                            ]
                        ),

                    "PRECIO_UNITARIO":
                        float(
                            detalle[
                                "precio_unitario"
                            ]
                        ),

                    "SUBTOTAL":
                        float(
                            detalle[
                                "subtotal"
                            ]
                        ),

                    "ID_COMPRA":
                        detalle[
                            "id_compra"
                        ],

                    "FECHA_REGISTRO":
                        fecha_registro,
                },
            )

            siguiente_id_detalle += 1

    except Exception as error:

        workbook.close()

        return (
            False,
            f"No se pudo preparar la factura: {error}",
        )

    # ========================================================
    # GUARDADO SEGURO
    # ========================================================

    archivo_temporal = (
        EXCEL_PATH.parent
        / "Control_Electropart_temp.xlsx"
    )

    try:

        # Guardamos primero en un archivo temporal.
        workbook.save(
            archivo_temporal
        )

        workbook.close()

        # Solamente cuando el archivo temporal quedó
        # correctamente creado sustituimos el original.
        os.replace(
            archivo_temporal,
            EXCEL_PATH,
        )

    except PermissionError:

        workbook.close()

        if archivo_temporal.exists():
            archivo_temporal.unlink()

        return (
            False,
            "No se pudo guardar la factura. "
            "Cierre Control_Electropart.xlsx e inténtelo nuevamente.",
        )

    except Exception as error:

        workbook.close()

        if archivo_temporal.exists():
            archivo_temporal.unlink()

        return (
            False,
            f"No se pudo guardar la factura: {error}",
        )

    # ========================================================
    # RESULTADO
    # ========================================================

    return (
        True,
        {
            "mensaje": "Factura registrada correctamente.",
            "id_factura": id_factura,
            "subtotal": calculos[
                "subtotal"
            ],
            "iva": calculos[
                "iva"
            ],
            "total": calculos[
                "total"
            ],
            "fecha_vencimiento":
                fecha_vencimiento,
        },
    )