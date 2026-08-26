from datetime import datetime
import re
import unicodedata

from openpyxl import load_workbook

from services.excel_service import (
    EXCEL_PATH,
    inicializar_estructura,
    obtener_columnas,
)


# ============================================================
# CONFIGURACIÓN
# ============================================================

HOJA_PRODUCTOS = "Productos"


CATEGORIAS_PRODUCTO = {
    "EQUIPO",
    "REPUESTO",
    "MATERIAL",
    "OTRO",
}


# ============================================================
# PREFIJOS PARA CÓDIGOS AUTOMÁTICOS
# ============================================================

PREFIJOS_CATEGORIA = {
    "REPUESTO": "REP",
    "EQUIPO": "EQU",
    "MATERIAL": "MAT",
    "OTRO": "OTR",
}


# ============================================================
# NORMALIZAR TEXTO
# ============================================================

def _normalizar_texto(texto):
    """
    Convierte un texto a una forma comparable.

    Ejemplo:

        "Termóstato Digital"
                ↓
        "termostato digital"
    """

    texto = str(
        texto or ""
    ).strip().casefold()

    texto = unicodedata.normalize(
        "NFD",
        texto,
    )

    texto = "".join(
        caracter
        for caracter in texto
        if unicodedata.category(
            caracter
        ) != "Mn"
    )

    return texto


# ============================================================
# SIGUIENTE ID
# ============================================================

def _obtener_siguiente_id(
    hoja,
    columna_id,
):
    """
    Devuelve el siguiente ID interno disponible.
    """

    mayor_id = 0

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        valor = hoja.cell(
            fila,
            columna_id,
        ).value

        if valor is None:
            continue

        try:

            valor = int(valor)

        except (
            TypeError,
            ValueError,
        ):
            continue

        if valor > mayor_id:
            mayor_id = valor

    return mayor_id + 1


# ============================================================
# GENERAR CÓDIGO AUTOMÁTICO
# ============================================================

def _generar_codigo_producto(
    hoja,
    columnas,
    categoria,
):
    """
    Genera automáticamente el código interno.

    Ejemplos:

        REPUESTO
            ↓
        REP-0001
        REP-0002

        EQUIPO
            ↓
        EQU-0001

        MATERIAL
            ↓
        MAT-0001
    """

    categoria = str(
        categoria or ""
    ).strip().upper()

    prefijo = PREFIJOS_CATEGORIA.get(
        categoria
    )

    if prefijo is None:

        raise ValueError(
            "No existe un prefijo para "
            "la categoría seleccionada."
        )

    # --------------------------------------------------------
    # Ejemplo de patrón:
    #
    # REP-0001
    # REP-0025
    # REP-1530
    # --------------------------------------------------------

    patron = re.compile(
        rf"^{re.escape(prefijo)}-(\d+)$",
        re.IGNORECASE,
    )

    mayor_numero = 0

    codigos_existentes = set()

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        codigo = str(
            hoja.cell(
                fila,
                columnas["CODIGO_INTERNO"],
            ).value
            or ""
        ).strip().upper()

        if codigo == "":
            continue

        codigos_existentes.add(
            codigo
        )

        coincidencia = patron.match(
            codigo
        )

        if coincidencia is None:
            continue

        numero = int(
            coincidencia.group(1)
        )

        if numero > mayor_numero:
            mayor_numero = numero

    # --------------------------------------------------------
    # GENERAR SIGUIENTE
    # --------------------------------------------------------

    siguiente_numero = (
        mayor_numero + 1
    )

    codigo_nuevo = (
        f"{prefijo}-"
        f"{siguiente_numero:04d}"
    )

    # Seguridad adicional ante cualquier inconsistencia.
    while (
        codigo_nuevo
        in codigos_existentes
    ):

        siguiente_numero += 1

        codigo_nuevo = (
            f"{prefijo}-"
            f"{siguiente_numero:04d}"
        )

    return codigo_nuevo


# ============================================================
# LISTAR PRODUCTOS
# ============================================================

def listar_productos():
    """
    Devuelve todos los productos registrados.
    """

    inicializar_estructura()

    workbook = load_workbook(
        EXCEL_PATH,
        data_only=True,
    )

    hoja = workbook[
        HOJA_PRODUCTOS
    ]

    columnas = obtener_columnas(
        hoja
    )

    productos = []

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        producto_id = hoja.cell(
            fila,
            columnas["ID_PRODUCTO"],
        ).value

        nombre = hoja.cell(
            fila,
            columnas["NOMBRE"],
        ).value

        # Ignorar filas completamente vacías.
        if (
            producto_id is None
            and nombre is None
        ):
            continue

        activo = hoja.cell(
            fila,
            columnas["ACTIVO"],
        ).value

        if activo is None:
            activo = True

        producto = {

            "id":
                producto_id,

            "nombre":
                str(
                    nombre or ""
                ).strip(),

            "categoria":
                str(
                    hoja.cell(
                        fila,
                        columnas["CATEGORIA"],
                    ).value
                    or ""
                ).strip(),

            "marca":
                str(
                    hoja.cell(
                        fila,
                        columnas["MARCA"],
                    ).value
                    or ""
                ).strip(),

            "modelo":
                str(
                    hoja.cell(
                        fila,
                        columnas["MODELO"],
                    ).value
                    or ""
                ).strip(),

            "codigo_interno":
                str(
                    hoja.cell(
                        fila,
                        columnas[
                            "CODIGO_INTERNO"
                        ],
                    ).value
                    or ""
                ).strip(),

            "unidad":
                str(
                    hoja.cell(
                        fila,
                        columnas["UNIDAD"],
                    ).value
                    or ""
                ).strip(),

            "descripcion":
                str(
                    hoja.cell(
                        fila,
                        columnas[
                            "DESCRIPCION"
                        ],
                    ).value
                    or ""
                ).strip(),

            "activo":
                bool(activo),

            "fecha_registro":
                hoja.cell(
                    fila,
                    columnas[
                        "FECHA_REGISTRO"
                    ],
                ).value,
        }

        productos.append(
            producto
        )

    workbook.close()

    return productos


# ============================================================
# BUSCAR PRODUCTOS
# ============================================================

def buscar_productos(
    texto,
    limite=15,
    solo_activos=True,
):
    """
    Busca por:

    - código;
    - nombre;
    - marca;
    - modelo.
    """

    texto = _normalizar_texto(
        texto
    )

    productos = listar_productos()

    coincidencias_inicio = []
    coincidencias_parciales = []

    for producto in productos:

        if (
            solo_activos
            and not producto["activo"]
        ):
            continue

        campos = [
            _normalizar_texto(
                producto["codigo_interno"]
            ),
            _normalizar_texto(
                producto["nombre"]
            ),
            _normalizar_texto(
                producto["marca"]
            ),
            _normalizar_texto(
                producto["modelo"]
            ),
        ]

        if texto == "":

            coincidencias_inicio.append(
                producto
            )

            continue

        # Prioridad:
        # el campo empieza por el texto buscado.
        if any(
            campo.startswith(texto)
            for campo in campos
            if campo
        ):

            coincidencias_inicio.append(
                producto
            )

            continue

        # Segunda prioridad:
        # contiene el texto.
        if any(
            texto in campo
            for campo in campos
            if campo
        ):

            coincidencias_parciales.append(
                producto
            )

    resultados = (
        coincidencias_inicio
        + coincidencias_parciales
    )

    return resultados[
        :limite
    ]


# ============================================================
# OBTENER PRODUCTO POR ID
# ============================================================

def obtener_producto_por_id(
    producto_id,
):
    """
    Busca un producto mediante su ID interno.
    """

    for producto in listar_productos():

        if producto["id"] == producto_id:

            return producto

    return None


# ============================================================
# REGISTRAR PRODUCTO
# ============================================================

def registrar_producto(
    nombre,
    categoria,
    marca="",
    modelo="",
    unidad="UNIDAD",
    descripcion="",
    activo=True,
):
    """
    Registra un nuevo producto.

    El código interno se genera automáticamente.

    El usuario NO proporciona CODIGO_INTERNO.
    """

    # ========================================================
    # LIMPIAR DATOS
    # ========================================================

    nombre = str(
        nombre or ""
    ).strip()

    categoria = str(
        categoria or ""
    ).strip().upper()

    marca = str(
        marca or ""
    ).strip()

    modelo = str(
        modelo or ""
    ).strip()

    unidad = str(
        unidad or "UNIDAD"
    ).strip().upper()

    descripcion = str(
        descripcion or ""
    ).strip()

    # ========================================================
    # VALIDACIONES
    # ========================================================

    if nombre == "":

        return (
            False,
            "El nombre del producto es obligatorio.",
        )

    if categoria not in CATEGORIAS_PRODUCTO:

        return (
            False,
            "La categoría del producto no es válida.",
        )

    if unidad == "":
        unidad = "UNIDAD"

    # ========================================================
    # ABRIR EXCEL
    # ========================================================

    inicializar_estructura()

    try:

        workbook = load_workbook(
            EXCEL_PATH
        )

    except PermissionError:

        return (
            False,
            "No se puede abrir el archivo Excel. "
            "Compruebe que esté cerrado.",
        )

    hoja = workbook[
        HOJA_PRODUCTOS
    ]

    columnas = obtener_columnas(
        hoja
    )

    # ========================================================
    # NORMALIZAR PARA DUPLICADOS
    # ========================================================

    nombre_normalizado = (
        _normalizar_texto(
            nombre
        )
    )

    marca_normalizada = (
        _normalizar_texto(
            marca
        )
    )

    modelo_normalizado = (
        _normalizar_texto(
            modelo
        )
    )

    # ========================================================
    # COMPROBAR DUPLICADOS
    # ========================================================

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        nombre_existente = str(
            hoja.cell(
                fila,
                columnas["NOMBRE"],
            ).value
            or ""
        ).strip()

        marca_existente = str(
            hoja.cell(
                fila,
                columnas["MARCA"],
            ).value
            or ""
        ).strip()

        modelo_existente = str(
            hoja.cell(
                fila,
                columnas["MODELO"],
            ).value
            or ""
        ).strip()

        if (
            _normalizar_texto(
                nombre_existente
            ) == nombre_normalizado
            and
            _normalizar_texto(
                marca_existente
            ) == marca_normalizada
            and
            _normalizar_texto(
                modelo_existente
            ) == modelo_normalizado
        ):

            workbook.close()

            return (
                False,
                "Ese producto ya se encuentra registrado.",
            )

    # ========================================================
    # GENERAR ID
    # ========================================================

    producto_id = (
        _obtener_siguiente_id(
            hoja,
            columnas[
                "ID_PRODUCTO"
            ],
        )
    )

    # ========================================================
    # GENERAR CÓDIGO AUTOMÁTICO
    # ========================================================

    try:

        codigo_interno = (
            _generar_codigo_producto(
                hoja=hoja,
                columnas=columnas,
                categoria=categoria,
            )
        )

    except ValueError as error:

        workbook.close()

        return (
            False,
            str(error),
        )

    # ========================================================
    # PREPARAR FILA
    # ========================================================

    nueva_fila = (
        hoja.max_row + 1
    )

    fecha_registro = (
        datetime.now()
    )

    datos = {

        "ID_PRODUCTO":
            producto_id,

        "NOMBRE":
            nombre,

        "CATEGORIA":
            categoria,

        "MARCA":
            marca,

        "MODELO":
            modelo,

        "CODIGO_INTERNO":
            codigo_interno,

        "UNIDAD":
            unidad,

        "DESCRIPCION":
            descripcion,

        "ACTIVO":
            bool(activo),

        "FECHA_REGISTRO":
            fecha_registro,
    }

    # ========================================================
    # ESCRIBIR
    # ========================================================

    for nombre_columna, valor in (
        datos.items()
    ):

        hoja.cell(
            nueva_fila,
            columnas[
                nombre_columna
            ],
            valor,
        )

    # ========================================================
    # GUARDAR
    # ========================================================

    try:

        workbook.save(
            EXCEL_PATH
        )

    except PermissionError:

        workbook.close()

        return (
            False,
            "No se pudo guardar el producto. "
            "Cierre Control_Electropart.xlsx.",
        )

    workbook.close()

    return (
        True,
        {
            "mensaje":
                "Producto registrado correctamente.",

            "id_producto":
                producto_id,

            "codigo_interno":
                codigo_interno,
        },
    )


# ============================================================
# ACTUALIZAR PRODUCTO
# ============================================================

def actualizar_producto(
    producto_id,
    nombre,
    categoria,
    marca="",
    modelo="",
    unidad="UNIDAD",
    descripcion="",
    activo=True,
):
    """
    Actualiza un producto.

    IMPORTANTE:

    CODIGO_INTERNO no se modifica.

    Aunque posteriormente se cambie la categoría,
    el código histórico permanece igual.
    """

    # ========================================================
    # LIMPIAR
    # ========================================================

    nombre = str(
        nombre or ""
    ).strip()

    categoria = str(
        categoria or ""
    ).strip().upper()

    marca = str(
        marca or ""
    ).strip()

    modelo = str(
        modelo or ""
    ).strip()

    unidad = str(
        unidad or "UNIDAD"
    ).strip().upper()

    descripcion = str(
        descripcion or ""
    ).strip()

    # ========================================================
    # VALIDACIONES
    # ========================================================

    if nombre == "":

        return (
            False,
            "El nombre del producto es obligatorio.",
        )

    if categoria not in CATEGORIAS_PRODUCTO:

        return (
            False,
            "La categoría del producto no es válida.",
        )

    if unidad == "":
        unidad = "UNIDAD"

    # ========================================================
    # ABRIR EXCEL
    # ========================================================

    inicializar_estructura()

    try:

        workbook = load_workbook(
            EXCEL_PATH
        )

    except PermissionError:

        return (
            False,
            "No se puede abrir el archivo Excel. "
            "Compruebe que esté cerrado.",
        )

    hoja = workbook[
        HOJA_PRODUCTOS
    ]

    columnas = obtener_columnas(
        hoja
    )

    fila_objetivo = None

    nombre_normalizado = (
        _normalizar_texto(
            nombre
        )
    )

    marca_normalizada = (
        _normalizar_texto(
            marca
        )
    )

    modelo_normalizado = (
        _normalizar_texto(
            modelo
        )
    )

    codigo_actual = ""

    # ========================================================
    # BUSCAR Y COMPROBAR DUPLICADOS
    # ========================================================

    for fila in range(
        2,
        hoja.max_row + 1,
    ):

        id_existente = hoja.cell(
            fila,
            columnas[
                "ID_PRODUCTO"
            ],
        ).value

        nombre_existente = str(
            hoja.cell(
                fila,
                columnas["NOMBRE"],
            ).value
            or ""
        ).strip()

        marca_existente = str(
            hoja.cell(
                fila,
                columnas["MARCA"],
            ).value
            or ""
        ).strip()

        modelo_existente = str(
            hoja.cell(
                fila,
                columnas["MODELO"],
            ).value
            or ""
        ).strip()

        # ----------------------------------------------------
        # PRODUCTO QUE ESTAMOS EDITANDO
        # ----------------------------------------------------

        if id_existente == producto_id:

            fila_objetivo = fila

            codigo_actual = str(
                hoja.cell(
                    fila,
                    columnas[
                        "CODIGO_INTERNO"
                    ],
                ).value
                or ""
            ).strip()

            continue

        # ----------------------------------------------------
        # DUPLICADO
        # ----------------------------------------------------

        if (
            _normalizar_texto(
                nombre_existente
            ) == nombre_normalizado
            and
            _normalizar_texto(
                marca_existente
            ) == marca_normalizada
            and
            _normalizar_texto(
                modelo_existente
            ) == modelo_normalizado
        ):

            workbook.close()

            return (
                False,
                "Otro producto ya tiene "
                "el mismo nombre, marca y modelo.",
            )

    # ========================================================
    # NO ENCONTRADO
    # ========================================================

    if fila_objetivo is None:

        workbook.close()

        return (
            False,
            "No se encontró el producto.",
        )

    # ========================================================
    # ACTUALIZAR
    # ========================================================

    datos = {

        "NOMBRE":
            nombre,

        "CATEGORIA":
            categoria,

        "MARCA":
            marca,

        "MODELO":
            modelo,

        "UNIDAD":
            unidad,

        "DESCRIPCION":
            descripcion,

        "ACTIVO":
            bool(activo),
    }

    # Fíjate que CODIGO_INTERNO NO está aquí.

    for nombre_columna, valor in (
        datos.items()
    ):

        hoja.cell(
            fila_objetivo,
            columnas[
                nombre_columna
            ],
            valor,
        )

    # ========================================================
    # GUARDAR
    # ========================================================

    try:

        workbook.save(
            EXCEL_PATH
        )

    except PermissionError:

        workbook.close()

        return (
            False,
            "No se pudo actualizar el producto. "
            "Cierre Control_Electropart.xlsx.",
        )

    workbook.close()

    return (
        True,
        {
            "mensaje":
                "Producto actualizado correctamente.",

            "codigo_interno":
                codigo_actual,
        },
    )